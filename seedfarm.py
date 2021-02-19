#!/usr/bin/env python.
import openpyxl, re, os, codecs

ugly_pattern = "Good Seed: (\w*).+?(?=:): (\d+).+?(?=:): (\d+).+?(?=:): (\d+).+?(?=:): (\d+).+?(?=:): (\w*).+?(?=:): (\w*)"

#Path to log
log_path = r"C:\Users\<user_name>\Documents\Battle Brothers\log.html"

stats = [
		"Matk",
		"Mdef",
		"Res",
		"Fat"]

old_seeds = []
new_seeds = []

#simplest class in the world. bunch of variables.
class goodSeed:
	seed = ""
	res = 0
	fat = 0
	msk = 0
	md = 0
	trait_one = ""
	trait_two = ""

	def __init__(self, seed):
			self.seed = seed

#scrape text for seed, stats and traits
def getData(text):
	
	for(seed, msk, md, res, fat, trait_one, trait_two) in re.findall(ugly_pattern, text):
		if seed not in old_seeds:
			s = goodSeed(seed)
			s.msk = msk
			s.md = md
			s.res = res
			s.fat = fat
			s.trait_one = trait_one
			s.trait_two = trait_two
			new_seeds.append(s)
		

def main():
	wb = openpyxl.load_workbook(filename = "LW seed farm.xlsx")
	ws = wb.active
	
	row = 0
	
	for cell in ws["A"]:
		if cell.value is None:
			break
		elif cell.value != "Seeds":
			old_seeds.append(cell.value)
			row +=1
	
	file = codecs.open(log_path, "r")
	text = file.read()
	file.close()
	
	getData(text)
	
	for s in new_seeds:
		ws.cell(row=row, column=1).value = s.seed
		ws.cell(row=row, column=2).value = int(s.msk)
		ws.cell(row=row, column=3).value = int(s.md)
		ws.cell(row=row, column=4).value = int(s.res)
		ws.cell(row=row, column=5).value = int(s.fat)
		ws.cell(row=row, column=6).value = s.trait_one
		ws.cell(row=row, column=7).value = s.trait_two
		row +=1
		
	wb.save("LW seed farm.xlsx")
	
if __name__ == '__main__':
    main()